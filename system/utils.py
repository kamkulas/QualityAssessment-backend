import openpyxl
from decimal import Decimal


class Reader:
    def __init__(self, system):
        self.system = system

    @staticmethod
    def format_value(value):
        if isinstance(value, str):
            return Decimal(value.replace(',', '.'))
        return Decimal(value)

    def read(self):
        path = self.system.filename
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        if worksheet.max_column % 4 != 0:
            return {'ok': False, 'message': 'Wrong number of columns.'}

        loops_number = worksheet.max_column // 4
        rows_number = worksheet.max_row
        loops = []

        required_columns = {'mv', 'ma', 'spa', 'cv'}
        for i in range(0, loops_number):
            columns = []
            loop = {}
            for j in range(1, 5):
                column = worksheet.cell(row=1, column=j+i*4).value
                columns.append(column)
                if column.split(':')[1] in required_columns:
                    loop[column.split(':')[1]] = [
                        self.format_value(worksheet.cell(row=iterator, column=j+i*4).value) for iterator in
                        range(3, rows_number+1)
                    ]
            loop_name = [column.split(':')[0] for column in columns]
            if len(set(loop_name)) > 1:
                return {'ok': False, 'message': 'Incorrect column names!'}
            loop['name'] = loop_name[0]
            loops.append(loop)

        return {'ok': True, 'loops': loops}